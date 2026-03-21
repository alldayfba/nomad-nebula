from __future__ import annotations
"""
Wholesale Manifest Processor — TA's Wholesale Search equivalent.
Upload a supplier CSV/Excel, get back every profitable ASIN with full profitability calc.

Usage:
    python execution/wholesale_manifest.py --file suppliers/price_list.csv --min-roi 30

CSV format (flexible column detection):
    Required: product name OR UPC (or both)
    Required: cost/price column
    Optional: ASIN, brand, category, quantity
"""
import os
import sys
import csv
import json
import logging
from typing import Optional

logger = logging.getLogger(__name__)
sys.path.insert(0, os.path.dirname(__file__))


def detect_columns(headers: list[str]) -> dict[str, int]:
    """Auto-detect column positions from header names."""
    col_map = {}
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if any(k in h_lower for k in ["upc", "barcode", "ean", "isbn", "gtin"]):
            col_map.setdefault("upc", i)
        elif any(k in h_lower for k in ["asin"]):
            col_map.setdefault("asin", i)
        elif any(k in h_lower for k in ["title", "name", "description", "product", "item"]):
            col_map.setdefault("name", i)
        elif any(k in h_lower for k in ["cost", "price", "wholesale", "unit cost", "your price"]):
            col_map.setdefault("cost", i)
        elif any(k in h_lower for k in ["brand", "manufacturer", "mfg"]):
            col_map.setdefault("brand", i)
        elif any(k in h_lower for k in ["qty", "quantity", "pack", "count", "units"]):
            col_map.setdefault("qty", i)
    return col_map


def load_manifest(filepath: str) -> list[dict]:
    """Load CSV or Excel manifest file."""
    rows = []

    if filepath.endswith((".xlsx", ".xls")):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            col_map = detect_columns(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(_build_row(row, col_map))
        except ImportError:
            logger.error("openpyxl required for Excel files: pip install openpyxl")
            return []
    else:
        # CSV
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            headers = next(reader)
            col_map = detect_columns(headers)
            for row in reader:
                if not any(row):
                    continue
                rows.append(_build_row(row, col_map))

    logger.info(f"Loaded {len(rows)} rows from manifest")
    return [r for r in rows if r.get("cost", 0) > 0]  # Filter invalid rows


def _build_row(row, col_map: dict) -> dict:
    """Build a normalized row dict from raw row data."""
    def get_val(key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return str(val).strip() if val is not None else None

    cost_str = get_val("cost") or "0"
    try:
        cost = float(cost_str.replace("$", "").replace(",", ""))
    except (ValueError, AttributeError):
        cost = 0

    return {
        "upc": get_val("upc"),
        "asin": get_val("asin"),
        "name": get_val("name") or "",
        "cost": cost,
        "brand": get_val("brand") or "",
        "qty": int(get_val("qty") or 1),
    }


def process_manifest(
    filepath: str,
    min_roi: float = 30.0,
    min_profit: float = 3.00,
    output_csv: Optional[str] = None,
) -> list[dict]:
    """
    Process a wholesale manifest and return profitable products.

    Returns list of dicts with full profitability data, sorted by ROI descending.
    """
    from dotenv import load_dotenv
    load_dotenv()

    rows = load_manifest(filepath)
    if not rows:
        logger.error("No valid rows found in manifest")
        return []

    # Try to import matching infrastructure
    try:
        from keepa_client import KeepaClient
        keepa = KeepaClient(os.getenv("KEEPA_API_KEY"))
    except Exception as e:
        logger.error(f"Keepa required for manifest processing: {e}")
        return []

    try:
        from calculate_fba_profitability import calculate_product_profitability
    except ImportError as e:
        logger.error(f"Profitability calc not available: {e}")
        return []

    results = []
    logger.info(f"Processing {len(rows)} manifest rows...")

    for i, row in enumerate(rows):
        if i % 50 == 0 and i > 0:
            logger.info(f"Progress: {i}/{len(rows)} ({len(results)} profitable found so far)")

        asin = row.get("asin")
        upc = row.get("upc")
        name = row.get("name", "")
        cost = row["cost"]

        # Step 1: Find the Amazon ASIN
        if not asin:
            if upc:
                # UPC to ASIN via Keepa
                try:
                    matches = keepa.search_by_upc(upc)
                    if matches:
                        asin = matches[0].get("asin")
                except Exception:
                    pass

            if not asin and name:
                # Title search via Keepa
                try:
                    matches = keepa.search_products(name, limit=1)
                    if matches:
                        asin = matches[0].get("asin")
                except Exception:
                    pass

        if not asin:
            continue

        # Step 2: Get Amazon data
        try:
            products = keepa.get_products([asin])
            if not products:
                continue
            product = products[0]
        except Exception:
            continue

        # Step 3: Calculate profitability
        sell_price = product.get("fba_price") or product.get("buy_box_price") or product.get("amazon_price", 0)
        if not sell_price or sell_price <= 0:
            continue

        try:
            prof = calculate_product_profitability(
                buy_cost=cost,
                sell_price=sell_price,
                weight_lbs=product.get("weight_lbs", 0.5),
                category=product.get("category", ""),
                fba_seller_count=product.get("fba_seller_count", 0),
                bsr=product.get("bsr", 0),
            )
        except Exception as e:
            logger.debug(f"Profitability calc failed for {asin}: {e}")
            continue

        roi = prof.get("roi_percent", 0)
        profit = prof.get("profit_per_unit", 0)

        if roi >= min_roi and profit >= min_profit:
            results.append({
                "asin": asin,
                "amazon_url": f"https://amazon.com/dp/{asin}",
                "name": product.get("title", name)[:80],
                "buy_cost": cost,
                "sell_price": sell_price,
                "profit": round(profit, 2),
                "roi": round(roi, 1),
                "max_cost": prof.get("max_cost", 0),
                "bsr": product.get("bsr", 0),
                "fba_sellers": product.get("fba_seller_count", 0),
                "verdict": prof.get("verdict", "MAYBE"),
                "category": product.get("category", ""),
                "source_upc": upc,
                "source_name": name,
            })

    # Sort by ROI
    results.sort(key=lambda x: x["roi"], reverse=True)

    # Optionally export CSV
    if output_csv and results:
        with open(output_csv, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)
        logger.info(f"Exported {len(results)} results to {output_csv}")

    logger.info(f"Manifest processing complete: {len(results)}/{len(rows)} profitable products found")
    return results


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Process wholesale supplier manifest")
    parser.add_argument("--file", required=True, help="Path to CSV or Excel manifest")
    parser.add_argument("--min-roi", type=float, default=30.0)
    parser.add_argument("--min-profit", type=float, default=3.00)
    parser.add_argument("--output", help="Output CSV path for results")
    args = parser.parse_args()

    results = process_manifest(args.file, args.min_roi, args.min_profit, args.output)
    if not args.output:
        print(json.dumps(results[:10], indent=2))
        if len(results) > 10:
            print(f"... and {len(results)-10} more. Use --output to export all.")
